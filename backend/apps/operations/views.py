from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import LogTravail, Site, TacheCatalogue
from .serializers import LogTravailSerializer, SiteSerializer, TacheCatalogueSerializer


def _get_payroll_data(site_id=None, mois=None, annee=None):
    """Helper : retourne les données groupées de paie à la tâche."""
    if mois is None:
        mois = timezone.now().month
    if annee is None:
        annee = timezone.now().year

    logs = LogTravail.objects.select_related("employe", "tache", "site")
    if site_id:
        logs = logs.filter(site_id=site_id)
    logs = logs.filter(date__month=mois, date__year=annee)

    grouped = {}
    for log in logs:
        key = (log.employe_id, log.tache_id)
        if key not in grouped:
            grouped[key] = {
                "employe_id": log.employe_id,
                "employe_nom": log.employe.nom_complet,
                "employe_code": log.employe.code,
                "employe_telephone": log.employe.telephone,
                "tache_id": log.tache_id,
                "tache_libelle": log.tache.libelle,
                "tache_unite": log.tache.unite_label,
                "tarif": float(log.tache.tarif_reference or 0),
                "site_id": log.site_id,
                "site_nom": log.site.nom,
                "quantite_totale": 0,
                "prime": 0,
                "montant": 0,
            }
        g = grouped[key]
        g["quantite_totale"] += float(log.objectif_realise)
        g["prime"] += float(log.prime or 0)
        g["montant"] = g["quantite_totale"] * g["tarif"] + g["prime"]

    results = sorted(grouped.values(), key=lambda x: x["employe_nom"])
    total = sum(r["montant"] for r in results)
    site_nom = logs[0].site.nom if site_id and results else None
    return results, total, site_nom, mois, annee


def _build_payroll_excel(results, total, site_nom, mois, annee):
    """Construit le fichier Excel au format EKO."""
    wb = Workbook()
    ws = wb.active
    ws.title = site_nom[:31] if site_nom else "Paie"

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F8F53", end_color="1F8F53", fill_type="solid"
    )
    title_font = Font(name="Calibri", size=14, bold=True, color="1F8F53")
    total_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = "#,##0"

    # Title
    mois_fr = [
        "",
        "Janvier",
        "Février",
        "Mars",
        "Avril",
        "Mai",
        "Juin",
        "Juillet",
        "Août",
        "Septembre",
        "Octobre",
        "Novembre",
        "Décembre",
    ]
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"ETAT DE PAIE À LA TÂCHE — {site_nom or 'Tous'} — {mois_fr[mois]} {annee}"
    )
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "N°",
        "NOM & PRENOMS",
        "TÂCHE",
        "QUANTITÉ",
        "PU (FCFA)",
        "PRIME",
        "MONTANT (FCFA)",
        "CONTACT",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data
    for i, r in enumerate(results, 1):
        row = i + 3
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=r["employe_nom"]).border = thin_border
        ws.cell(row=row, column=3, value=r["tache_libelle"]).border = thin_border
        qte = ws.cell(row=row, column=4, value=r["quantite_totale"])
        qte.border = thin_border
        ws.cell(row=row, column=5, value=r["tarif"]).border = thin_border
        ws.cell(row=row, column=5).number_format = money_fmt
        ws.cell(row=row, column=6, value=r.get("prime", 0)).border = thin_border
        ws.cell(row=row, column=6).number_format = money_fmt
        montant = ws.cell(row=row, column=7, value=r["montant"])
        montant.border = thin_border
        montant.number_format = money_fmt
        ws.cell(
            row=row, column=8, value=r.get("employe_telephone", "")
        ).border = thin_border

    # Total row
    total_row = len(results) + 4
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        ws.cell(row=total_row, column=c).border = thin_border
    total_cell = ws.cell(row=total_row, column=7, value=total)
    total_cell.font = total_font
    total_cell.number_format = money_fmt
    total_cell.border = thin_border
    ws.cell(row=total_row, column=7).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    return wb


class SiteViewSet(viewsets.ModelViewSet):
    queryset = Site.objects.all()
    serializer_class = SiteSerializer
    filterset_fields = ["type_site", "actif"]
    search_fields = ["code", "nom"]


class TacheCatalogueViewSet(viewsets.ModelViewSet):
    queryset = TacheCatalogue.objects.all()
    serializer_class = TacheCatalogueSerializer
    filterset_fields = ["type_objectif", "actif"]
    search_fields = ["code", "libelle"]


class LogTravailViewSet(viewsets.ModelViewSet):
    queryset = LogTravail.objects.select_related("employe", "site", "tache")
    serializer_class = LogTravailSerializer
    filterset_fields = ["employe", "site", "tache", "date"]
    search_fields = ["employe__nom", "site__nom", "tache__libelle", "notes"]

    @action(detail=False, methods=["get"])
    def task_payroll(self, request):
        """Paie à la tâche : regroupé par employé × tâche avec montant calculé."""
        site_id = request.query_params.get("site")
        mois = request.query_params.get("mois")
        annee = request.query_params.get("annee")
        if mois:
            mois = int(mois)
        if annee:
            annee = int(annee)

        results, total, site_nom, mois, annee = _get_payroll_data(site_id, mois, annee)

        return Response(
            {
                "mois": mois,
                "annee": annee,
                "site_nom": site_nom,
                "lignes": results,
                "total": total,
            }
        )

    @action(detail=False, methods=["get"])
    def export_task_payroll(self, request):
        """Export Excel de la paie à la tâche."""
        site_id = request.query_params.get("site")
        mois = request.query_params.get("mois")
        annee = request.query_params.get("annee")
        if mois:
            mois = int(mois)
        if annee:
            annee = int(annee)

        results, total, site_nom, mois, annee = _get_payroll_data(site_id, mois, annee)
        wb = _build_payroll_excel(results, total, site_nom, mois, annee)

        safe_name = site_nom.replace(" ", "_") if site_nom else "tous"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="paie_tache_{safe_name}_{mois}_{annee}.xlsx"'
        )
        wb.save(response)
        return response

    @action(detail=True, methods=["post"])
    def marquer_paye(self, request, pk=None):
        """Marque un log de travail comme payé et crée un Paiement."""
        log = self.get_object()
        if log.paye_le:
            return Response({"error": "Déjà payé"}, status=400)

        log.paye_le = request.data.get("paye_le") or timezone.now().date()
        log.save()

        # Créer un paiement
        from apps.rh.models import Paiement

        montant = float(log.objectif_realise) * float(
            log.tache.tarif_reference or 0
        ) + float(log.prime or 0)
        Paiement.objects.create(
            employe=log.employe,
            date=log.paye_le,
            montant=montant,
            mode=request.data.get("mode", "especes"),
            notes=f"Log #{log.id} — {log.tache.libelle}",
        )

        return Response(
            {"status": "payé", "paye_le": str(log.paye_le), "montant": montant}
        )
