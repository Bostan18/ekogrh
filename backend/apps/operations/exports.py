"""Exports Excel pour le module Opérations."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _get_payroll_data(logs_queryset):
    """Helper : retourne les données groupées de paie à la tâche."""
    grouped = {}
    for log in logs_queryset:
        key = (log.employe_id, log.tache_id)
        if key not in grouped:
            grouped[key] = {
                "employe_id": log.employe_id,
                "employe_nom": log.employe.nom_complet,
                "employe_code": log.employe.code,
                "employe_telephone": log.employe.telephone,
                "tache_id": log.tache_id,
                "tache_libelle": log.tache.libelle,
                "tache_unite": log.tache.unite_label,
                "tarif": float(log.tache.tarif_reference or 0),
                "seuil": float(log.tache.seuil or 0),
                "site_id": log.site_id,
                "site_nom": log.site.nom,
                "quantite_totale": 0,
                "prime": 0,
                "montant": 0,
            }
        g = grouped[key]
        g["quantite_totale"] += float(log.objectif_realise)
        g["prime"] += float(log.prime or 0)
        qte_payable = max(0, g["quantite_totale"] - g["seuil"])
        g["montant"] = qte_payable * g["tarif"] + g["prime"]

    results = sorted(grouped.values(), key=lambda x: x["employe_nom"])
    total = sum(r["montant"] for r in results)
    return results, total


from apps.core.constants import MOIS_FR


def build_payroll_excel(results, total, site_nom, mois, annee):
    """Construit le fichier Excel de paie à la tâche au format EKO."""
    wb = Workbook()
    ws = wb.active
    ws.title = (site_nom or "Paie")[:31]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F8F53", end_color="1F8F53", fill_type="solid"
    )
    title_font = Font(name="Calibri", size=14, bold=True, color="1F8F53")
    total_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = "#,##0"

    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"ETAT DE PAIE À LA TÂCHE — {site_nom or 'Tous'} — {MOIS_FR[mois]} {annee}"
    )
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "N°",
        "NOM & PRENOMS",
        "TÂCHE",
        "QUANTITÉ",
        "PU (FCFA)",
        "PRIME",
        "MONTANT (FCFA)",
        "CONTACT",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, 1):
        row = i + 3
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=r["employe_nom"]).border = thin_border
        ws.cell(row=row, column=3, value=r["tache_libelle"]).border = thin_border
        ws.cell(row=row, column=4, value=r["quantite_totale"]).border = thin_border
        c = ws.cell(row=row, column=5, value=r["tarif"])
        c.number_format = money_fmt
        c.border = thin_border
        c = ws.cell(row=row, column=6, value=r.get("prime", 0))
        c.number_format = money_fmt
        c.border = thin_border
        c = ws.cell(row=row, column=7, value=r["montant"])
        c.number_format = money_fmt
        c.border = thin_border
        ws.cell(
            row=row, column=8, value=r.get("employe_telephone", "")
        ).border = thin_border

    total_row = len(results) + 4
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).border = thin_border
    c = ws.cell(row=total_row, column=7, value=total)
    c.font = total_font
    c.number_format = money_fmt
    c.border = thin_border

    for col, w in zip("ABCDEFGH", [6, 30, 22, 12, 14, 12, 18, 18]):
        ws.column_dimensions[col].width = w

    return wb
