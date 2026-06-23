import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

VERT_EKO = "1F8F53"
VERT_CLAIR = "E8F5EE"

MOIS_FR = [
    "",
    "Janvier",
    "Février",
    "Mars",
    "Avril",
    "Mai",
    "Juin",
    "Juillet",
    "Août",
    "Septembre",
    "Octobre",
    "Novembre",
    "Décembre",
]


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def paie_excel(employes, presences_par_employe, mois, annee):
    """Génère la feuille de paie mensuelle."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Paie {MOIS_FR[mois]} {annee}"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    sub_fill = PatternFill("solid", fgColor=VERT_CLAIR)
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"FEUILLE DE PAIE — {MOIS_FR[mois].upper()} {annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Code", "Nom & Prénom", "Type", "Taux/Salaire (F)",
        "Jours présents", "Total à payer (F)", "Poste",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor="2D7A50")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 30

    TYPE_LABEL = {"cdi": "CDI", "cdd": "CDD", "journalier": "Journalier", "moo": "MOO"}

    total_global = 0
    row = 3
    for emp in employes:
        presences = presences_par_employe.get(emp.id, [])
        jours_presents = len([p for p in presences if p.present])

        if emp.type_contrat in ("cdi", "cdd"):
            taux = float(emp.salaire_mensuel or 0)
            total = taux
        else:
            taux = float(emp.taux_journalier or 0)
            total = sum(float(p.montant_du) for p in presences if p.present)

        total_global += total

        ws.cell(row=row, column=1, value=emp.code).border = _border()
        ws.cell(row=row, column=2, value=f"{emp.nom} {emp.prenom}").border = _border()
        ws.cell(row=row, column=3, value=TYPE_LABEL.get(emp.type_contrat, emp.type_contrat)).border = _border()
        cell = ws.cell(row=row, column=4, value=taux)
        cell.number_format = "#,##0"
        cell.border = _border()
        ws.cell(row=row, column=5, value=jours_presents if emp.type_contrat not in ("cdi", "cdd") else "—").border = _border()
        cell = ws.cell(row=row, column=6, value=total)
        cell.font = bold
        cell.number_format = "#,##0"
        cell.border = _border()
        ws.cell(row=row, column=7, value=emp.poste or "—").border = _border()
        row += 1

    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "MASSE SALARIALE TOTALE"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=VERT_CLAIR)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    cell = ws.cell(row=row, column=6, value=total_global)
    cell.font = Font(bold=True, size=11)
    cell.number_format = "#,##0"
    cell.fill = PatternFill("solid", fgColor=VERT_CLAIR)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def presences_excel(presences, mois, annee):
    """Génère la feuille de présences détaillée pour un mois."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Presences {MOIS_FR[mois]} {annee}"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    sub_fill = PatternFill("solid", fgColor=VERT_CLAIR)
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    red_font = Font(color="CC3333")
    green_font = Font(color="1A5C38")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"FEUILLE DE PRÉSENCES — {MOIS_FR[mois].upper()} {annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Date", "Code", "Employé", "Type", "Présent",
        "Heures", "Montant (F)", "Projet", "Site", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor="2D7A50")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 30

    TYPE_LABEL = {"cdi": "CDI", "cdd": "CDD", "journalier": "Journalier", "moo": "MOO"}

    total_presents = 0
    total_montant = 0.0
    row = 3
    for p in presences:
        ws.cell(row=row, column=1, value=str(p.date)).border = _border()
        ws.cell(row=row, column=2, value=p.employe.code).border = _border()
        ws.cell(row=row, column=3, value=f"{p.employe.nom} {p.employe.prenom}").border = _border()
        ws.cell(row=row, column=4, value=TYPE_LABEL.get(p.employe.type_contrat, p.employe.type_contrat)).border = _border()
        cell_present = ws.cell(row=row, column=5, value="✓ Présent" if p.present else "✗ Absent")
        cell_present.font = green_font if p.present else red_font
        cell_present.border = _border()
        ws.cell(row=row, column=6, value=float(p.heures_travaillees) if p.present else 0).border = _border()
        cell_montant = ws.cell(row=row, column=7, value=float(p.montant_du) if p.present else 0)
        cell_montant.number_format = "#,##0"
        cell_montant.border = _border()
        ws.cell(row=row, column=8, value=p.projet_ref or "—").border = _border()
        ws.cell(row=row, column=9, value=p.site_ref or "—").border = _border()
        ws.cell(row=row, column=10, value=p.notes or "—").border = _border()
        if p.present:
            total_presents += 1
            total_montant += float(p.montant_du)
        row += 1

    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = f"TOTAL — {total_presents} jour{['s', ''][total_presents == 1]} de présence"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=VERT_CLAIR)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    cell = ws.cell(row=row, column=7, value=total_montant)
    cell.font = Font(bold=True, size=11)
    cell.number_format = "#,##0"
    cell.fill = PatternFill("solid", fgColor=VERT_CLAIR)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def cnps_excel(bulletins, mois, annee):
    """Génère le fichier déclaratif CNPS pour un mois."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"CNPS {MOIS_FR[mois]} {annee}"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"DÉCLARATION CNPS — {MOIS_FR[mois].upper()} {annee}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Code", "Nom & Prénom", "Poste", "Brut (F)",
        "Part salariale 6,3%", "Part patronale", "Total cotisation", "Net",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor="2D7A50")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 30

    TAUX_SALARIAL = 0.063
    TAUX_PATRONAL = 0.16

    total_brut = 0.0
    total_salarial = 0.0
    total_patronal = 0.0
    row = 3
    for b in bulletins:
        brut = float(b.brut)
        part_salarial = round(brut * TAUX_SALARIAL, 0)
        part_patronal = round(brut * TAUX_PATRONAL, 0)
        total_cotisation = part_salarial + part_patronal
        total_brut += brut
        total_salarial += part_salarial
        total_patronal += part_patronal

        ws.cell(row=row, column=1, value=b.employe.code).border = _border()
        ws.cell(row=row, column=2, value=f"{b.employe.nom} {b.employe.prenom}").border = _border()
        ws.cell(row=row, column=3, value=b.employe.poste or "—").border = _border()
        c = ws.cell(row=row, column=4, value=brut)
        c.number_format = "#,##0"; c.border = _border()
        c = ws.cell(row=row, column=5, value=part_salarial)
        c.number_format = "#,##0"; c.border = _border()
        c = ws.cell(row=row, column=6, value=part_patronal)
        c.number_format = "#,##0"; c.border = _border()
        c = ws.cell(row=row, column=7, value=total_cotisation)
        c.number_format = "#,##0"; c.border = _border()
        c = ws.cell(row=row, column=8, value=float(b.net))
        c.number_format = "#,##0"; c.border = _border()
        row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "TOTAL"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=VERT_CLAIR)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    for col, val in [(4, total_brut), (5, total_salarial), (6, total_patronal), (7, total_salarial + total_patronal)]:
        c = ws.cell(row=row, column=col, value=round(val, 0))
        c.font = Font(bold=True, size=11)
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor=VERT_CLAIR)

    for col, w in zip("ABCDEFGH", [10, 28, 20, 14, 18, 16, 16, 14]):
        ws.column_dimensions[col].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def bordereau_journaliers_excel(employe_nom, presences, date_paiement):
    """Génère un bordereau de paiement pour un lot de journées d'un journalier."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bordereau {employe_nom[:20]}"

    header_fill = PatternFill("solid", fgColor=VERT_EKO)
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    ws.merge_cells("A1:E1")
    ws["A1"] = "EKO SARL — BORDEREAU DE PAIEMENT JOURNALIER"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Employé : {employe_nom}"
    ws["A2"].font = bold
    ws.merge_cells("D2:E2")
    ws["D2"] = f"Payé le : {date_paiement}"
    ws["D2"].font = bold
    ws["D2"].alignment = Alignment(horizontal="right")

    headers = ["Date", "Heures", "Projet", "Montant (F)", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor="2D7A50")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    total = 0.0
    row = 4
    for p in presences:
        montant = float(p.montant_du)
        total += montant
        ws.cell(row=row, column=1, value=str(p.date)).border = _border()
        ws.cell(row=row, column=2, value=f"{p.heures_travaillees}h").border = _border()
        ws.cell(row=row, column=3, value=p.projet_ref or "—").border = _border()
        c = ws.cell(row=row, column=4, value=montant)
        c.number_format = "#,##0"
        c.border = _border()
        ws.cell(row=row, column=5, value=p.notes or "—").border = _border()
        row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = f"TOTAL — {len(presences)} journée(s)"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=VERT_CLAIR)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    c = ws.cell(row=row, column=4, value=total)
    c.font = Font(bold=True, size=12)
    c.number_format = "#,##0"
    c.fill = PatternFill("solid", fgColor=VERT_CLAIR)

    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Signature employeur : ________________"
    ws[f"A{row}"].font = Font(italic=True, color="666666")
    ws.merge_cells(f"D{row}:E{row}")
    ws[f"D{row}"] = "Signature employé : ________________"
    ws[f"D{row}"].font = Font(italic=True, color="666666")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
